import openpyxl
from openpyxl import Workbook

def criar_planilha():
    # Crie uma nova planilha ou abra uma existente
    try:
        workbook = openpyxl.load_workbook('usuarios.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Usuário', 'Senha'])

    return workbook, sheet

def cadastrar_usuario():
    nome = input("Digite o nome de usuário: ")
    senha = input("Digite a senha: ")

    workbook, sheet = criar_planilha()
    sheet.append([nome, senha])
    workbook.save('usuarios.xlsx')
    print(f'Usuário {nome} cadastrado com sucesso!')

def fazer_login():
    nome = input("Digite o nome de usuário: ")
    senha = input("Digite a senha: ")

    workbook, sheet = criar_planilha()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == nome and row[1] == senha:
            print("Login bem-sucedido!")
            return
    print("Nome de usuário ou senha incorretos.")

def main():
    while True:
        print("1. Cadastrar usuário")
        print("2. Fazer login")
        print("3. Sair")
        escolha = input("Escolha uma opção: ")

        if escolha == '1':
            cadastrar_usuario()
        elif escolha == '2':
            fazer_login()
        elif escolha == '3':
            break
        else:
            print("Opção inválida. Tente novamente.")

if __name__ == "__main__":
    main()

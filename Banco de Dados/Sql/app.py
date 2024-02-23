import openpyxl
from openpyxl import Workbook
import os

def cls():
    os.system('cls')

def criar_planilha():
    # Crie uma nova planilha ou abra uma existente
    try:
        workbook = openpyxl.load_workbook(r'cadastro.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Nome', 'Cpf', 'Telefone', 'E-mail', 'endereço', "Data de Nascimento"])

    return workbook, sheet

def cadastrar_usuario():
    cls()
    # Solicitar nome de usuário
    nome = input("Digite o nome de usuário: ")

    # Solicitar CPF
    cpf = int(input("Digite o Cpf (Apenas Números): "))

    # Solicitar telefone
    telefone = (input("Digite o Telefone: "))

    # Solicitar e-mail
    e_mail = input("Digite o E-Mail:")

    # Solicitar endereço
    endereço = input("Digite o endereço:")

    # Solicitar data de nascimento
    data_de_nascimento = input("Digite a Data de nascimento(Ex: XX/XX/XXXX):")

    # Criar planilha para armazenar dados
    workbook, sheet = criar_planilha()

    # Adicionar dados na planilha
    sheet.append([nome, cpf, telefone, e_mail, endereço, data_de_nascimento])

    # Salvar planilha
    workbook.save('cadastro.xlsx')

    # Exibir mensagem de sucesso
    print(f'Usuário {nome} cadastrado com sucesso!')
    
def listar_cadastro():
    cls()
    i = 1
    Workbook, sheet = criar_planilha()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        
        print(f"------------------------ Usuario {i} -----------------------------\n")
        i = i+1
        nome = row[0]
        cpf  = row[1]
        tell  = row[2]
        mail  = row[3]
        endereco  = row[4]
        nascimento  = row[5]
        print(f"Nome: {nome}\nCPF: {cpf}\nTELEFONE: {tell}\nEMAIL: {mail}\nENDEREÇO: {endereco}\nNASCIMENTO: {nascimento}\n")

def mostrar_usuario():
    cls()
    listar_cadastro()
    input("Pressione enter para sair...")

def remover_registro():
    cls()
        # Criar planilha
    Workbook, sheet = criar_planilha()
    
        # Listar usuários
    listar_cadastro()
    
        # Solicitar linha a ser deletada
    linha_delet = int(input("Qual Usuario deseja Remover: "))
    
        # Ajustar linha para index da planilha
    linha_delet = linha_delet+1
    
        # Deletar linha
    sheet.delete_rows(linha_delet)
    
        # Salvar alterações
    Workbook.save('cadastro.xlsx')

    # Fechar arquivo da planilha
    Workbook.close()
    
    # Listar usuários novamente
    listar_cadastro()
    
def editar_registro():
    cls()
    Workbook, sheet = criar_planilha()
    listar_cadastro()
    edit_usuario = int(input("Qual o numero do usuario que você deseja editar:"))
    edit_usuario = edit_usuario+1
    cls()
    print("-----------------------------------------------------\n")
    print("1 - Nome\n2 - Cpf\n3 - Telefone\n4 - E-Mail\n5 - Endereço\n6 - Data de Nascimento")
    print("\n-----------------------------------------------------\n")
    opc = int(input("O que você deseja editar desse usuario:"))
    opcao = ""
    
    if 1 <= opc <= 6:
        if opc == 1:
            opcao = "Nome"
            new_value = input(f"Digite o novo valor para {opcao}: ")
        elif opc == 2:
            opcao = "Cpf"
            new_value = input(f"Digite o novo valor para {opcao}: ")
        elif opc == 3:
            opcao = "Telefone"
            new_value = input(f"Digite o novo valor para {opcao}: ")
        elif opc == 4:
            opcao = "E-mail"
            new_value = input(f"Digite o novo valor para {opcao}: ")
        elif opc == 5:
            opcao = "Endereço"
            new_value = input(f"Digite o novo valor para {opcao}: ")
        elif opc == 6:
            opcao = "Data de Nascimento"
            new_value = input(f"Digite o novo valor para {opcao}: ")
            
        sheet.cell(row=edit_usuario, column=opc, value=new_value)
        Workbook.save('cadastro.xlsx')
        Workbook.close()
        print(f"{opcao} do usuário {edit_usuario - 1} foi atualizado com sucesso para {new_value}")
    
    else:
        print("Opção inválida.")
    
while True: 
    cls()
    print(f"------------------------ JANELA PRINCIPAL -----------------------------\n")
    print("1 - Cadastrar Novo Usuario\n2 - Atualizar Dados\n3 - Listar cadastro\n4 - Excluir Usuario\n\n0 - Sair")
    print(f"-----------------------------------------------------------------------")
    choice = int(input("O que deseja fazer:"))
    
    match choice:   
        
        case 1: 
            cadastrar_usuario()
        case 2:
            editar_registro()
        case 3:
            mostrar_usuario()
        case 4:
            remover_registro()
        case 0:
            break        
        case _: 
            print("Opção Invalida")